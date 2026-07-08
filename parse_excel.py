import sys
import subprocess
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
import json

wb = openpyxl.load_workbook('Jadwal Depok Revisi.xlsx', data_only=True)
data = {}
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_data = []
    for row in sheet.iter_rows(values_only=True):
        sheet_data.append(row)
    data[sheet_name] = sheet_data

with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print("Excel data parsed and saved to excel_data.json")
